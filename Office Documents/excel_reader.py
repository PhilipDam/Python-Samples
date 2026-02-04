#!/usr/bin/env python3
"""
excel_reader.py

A heavily-commented, pragmatic Excel (.xlsx) reader built on top of openpyxl.

Goal
----
Load an .xlsx file and expose a small API so you can request:

- number of sheets
- list of sheet names
- selecting a sheet by index or by name
- left-most used column
- right-most used column
- top-most used row
- bottom-most used row
- value of any cell by [row, column] *within the used-area rectangle*
- typed accessors: string, optional bool, int, double
- first cell in row (within used area)
- last cell in row (within used area)
- used area size (#rows, #columns)

Important notes
---------------
- "Formatting information is irrelevant" and "only values are needed".
  Therefore, we compute the "used area" by scanning for actual *values*,
  not by relying on Excel formatting or openpyxl's max_row/max_column
  (which can be inflated by styles / previously-used cells).

Dependencies
------------
pip install openpyxl

Usage (demo CLI)
---------------
python3 excel_reader.py /path/to/file.xlsx
python3 excel_reader.py /path/to/file.xlsx "Sheet2"
python3 excel_reader.py /path/to/file.xlsx --sheet-index 0
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from typing import Any, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# Helper datatypes
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class UsedArea:
    """
    Represents the rectangle of used cells in a worksheet.

    Coordinates are Excel's native 1-based row/column indices.

    If a sheet is completely empty (no values), we represent that as:
        min_row = max_row = min_col = max_col = 0
    """
    min_row: int
    max_row: int
    min_col: int
    max_col: int

    @property
    def is_empty(self) -> bool:
        return self.min_row == 0

    @property
    def nrows(self) -> int:
        if self.is_empty:
            return 0
        return self.max_row - self.min_row + 1

    @property
    def ncols(self) -> int:
        if self.is_empty:
            return 0
        return self.max_col - self.min_col + 1


# ---------------------------------------------------------------------------
# Main reader class
# ---------------------------------------------------------------------------

class ExcelValueReader:
    """
    Loads an Excel workbook and provides convenient value-only access.

    Design decisions
    ----------------
    - We load with data_only=True so formula cells return their last cached result.
      (Excel stores calculated results; openpyxl does NOT evaluate formulas.)
    - We compute the used area by scanning for actual non-empty values.
      This avoids being fooled by formatting.
    - All "used-area-relative" coordinates are 1-based:
        used_row = 1 means the top-most used row (UsedArea.min_row)
        used_col = 1 means the left-most used column (UsedArea.min_col)
    """

    def __init__(self, filename: str):
        # data_only=True:
        #   if a cell has a formula like "=A1+B1", openpyxl returns the cached result
        #   (if present) instead of returning the formula text itself.
        # read_only=True:
        #   faster and less memory for big files (we only need values).
        self._wb = load_workbook(filename=filename, data_only=True, read_only=True)

        # Current worksheet (selected sheet)
        self._ws: Worksheet = self._wb.worksheets[0]

        # Cache used-area per sheet name to avoid rescanning repeatedly
        self._used_area_cache: dict[str, UsedArea] = {}

    # -------------------------
    # Workbook-level operations
    # -------------------------

    def sheet_count(self) -> int:
        """Return number of sheets in the workbook."""
        return len(self._wb.worksheets)

    def sheet_names(self) -> List[str]:
        """Return list of sheet names in workbook order."""
        return [ws.title for ws in self._wb.worksheets]

    def select_sheet_by_index(self, index: int) -> None:
        """
        Select a sheet by 0-based index (common in programming).
        Example: index=0 selects the first sheet.
        """
        self._ws = self._wb.worksheets[index]

    def select_sheet_by_name(self, name: str) -> None:
        """Select a sheet by exact name."""
        self._ws = self._wb[name]

    def selected_sheet_name(self) -> str:
        """Return the name of the currently selected sheet."""
        return self._ws.title

    # -------------------------
    # Used-area computations
    # -------------------------

    def used_area(self) -> UsedArea:
        """
        Compute (or retrieve cached) used area for the current worksheet.

        "Used" means: at least one cell in the rectangle has a value that is not None
        and not an all-whitespace string.

        Why not rely on ws.max_row/max_column?
        - Those can reflect formatting or historical edits rather than actual values.
        """
        title = self._ws.title
        if title in self._used_area_cache:
            return self._used_area_cache[title]

        min_row = 0
        max_row = 0
        min_col = 0
        max_col = 0

        # openpyxl iterates rows as tuples of Cell objects.
        # Using values_only=True yields raw values and is cheaper.
        for r_idx, row in enumerate(self._ws.iter_rows(values_only=True), start=1):
            # Quick skip: if the row has no values at all (all None), continue.
            # But note: empty strings "   " should be treated as empty too.
            row_has_value = False

            for c_idx, val in enumerate(row, start=1):
                if self._is_value_used(val):
                    row_has_value = True
                    if min_row == 0:
                        # first used cell we ever found
                        min_row = max_row = r_idx
                        min_col = max_col = c_idx
                    else:
                        # extend bounds
                        if r_idx < min_row:
                            min_row = r_idx
                        if r_idx > max_row:
                            max_row = r_idx
                        if c_idx < min_col:
                            min_col = c_idx
                        if c_idx > max_col:
                            max_col = c_idx

            # If row_has_value is False, we just continue scanning.
            # (We still need to scan all rows because used cells could appear later.)

        area = UsedArea(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)
        self._used_area_cache[title] = area
        return area

    @staticmethod
    def _is_value_used(val: Any) -> bool:
        """
        Decide if a cell value counts as 'used'.

        We treat:
        - None as unused
        - strings that are empty or whitespace-only as unused
        Everything else counts as used (numbers, bools, dates, text, etc.)
        """
        if val is None:
            return False
        if isinstance(val, str) and val.strip() == "":
            return False
        return True

    # Used-area boundary helpers (requested by user)

    def left_most_used_column(self) -> int:
        """Return left-most used column index (1-based). Returns 0 for empty sheet."""
        return self.used_area().min_col

    def right_most_used_column(self) -> int:
        """Return right-most used column index (1-based). Returns 0 for empty sheet."""
        return self.used_area().max_col

    def top_most_used_row(self) -> int:
        """Return top-most used row index (1-based). Returns 0 for empty sheet."""
        return self.used_area().min_row

    def bottom_most_used_row(self) -> int:
        """Return bottom-most used row index (1-based). Returns 0 for empty sheet."""
        return self.used_area().max_row

    def used_area_size(self) -> Tuple[int, int]:
        """Return (number_of_rows, number_of_columns) in used area."""
        area = self.used_area()
        return (area.nrows, area.ncols)

    # -------------------------
    # Value access (core)
    # -------------------------

    def get_value_by_used_coords(self, used_row: int, used_col: int) -> Any:
        """
        Get the raw value for a cell addressed *relative to the used area*.

        Example:
        - used_row=1, used_col=1 returns the value at (min_row, min_col)
        - used_row=area.nrows returns the bottom-most used row
        - used_col=area.ncols returns the right-most used column

        Coordinates are 1-based within the used area.

        Assumption from your request:
        "All row, column, cell information is always within the selected sheet"
        and "within the rectangular area described ..."
        So we do not add heavy bounds-checking beyond basic sanity.
        """
        area = self.used_area()
        if area.is_empty:
            raise ValueError("Selected sheet has no used cells (no values).")

        # Translate used-area-relative indices into Excel absolute indices
        abs_row = area.min_row + (used_row - 1)
        abs_col = area.min_col + (used_col - 1)

        return self.get_value_by_excel_coords(abs_row, abs_col)

    def get_value_by_excel_coords(self, row: int, col: int) -> Any:
        """
        Get raw value using Excel absolute coordinates (1-based row/col).

        This is a convenience method; your primary request was used-area-relative.
        """
        return self._ws.cell(row=row, column=col).value

    # -------------------------
    # Typed accessors (requested)
    # -------------------------

    def get_string(self, used_row: int, used_col: int) -> str:
        """
        Return the cell value as a string.

        - None -> ""
        - Everything else -> str(value)
        """
        v = self.get_value_by_used_coords(used_row, used_col)
        return "" if v is None else str(v)

    def get_bool(self, used_row: int, used_col: int) -> Optional[bool]:
        """
        Return the cell value as Optional[bool].

        Conversion rules:
        - If value is already bool => return it
        - If value is a string like "true/false/yes/no/1/0" (case-insensitive) => parse
        - If value is numeric 1/0 => convert
        - Otherwise => None
        """
        v = self.get_value_by_used_coords(used_row, used_col)

        if isinstance(v, bool):
            return v

        if isinstance(v, (int, float)):
            if v == 1:
                return True
            if v == 0:
                return False
            return None

        if isinstance(v, str):
            s = v.strip().lower()
            if s in ("true", "yes", "y", "1"):
                return True
            if s in ("false", "no", "n", "0"):
                return False
            return None

        return None

    def get_int(self, used_row: int, used_col: int) -> Optional[int]:
        """
        Return the cell value as Optional[int].

        Conversion rules:
        - int => int
        - float => only if it is integral (e.g. 42.0) => 42
        - bool => treat as int? (common trap) => we reject and return None
        - str => try int(stripped)
        - otherwise => None
        """
        v = self.get_value_by_used_coords(used_row, used_col)

        if isinstance(v, bool):
            return None

        if isinstance(v, int):
            return v

        if isinstance(v, float):
            if v.is_integer():
                return int(v)
            return None

        if isinstance(v, str):
            s = v.strip()
            if s == "":
                return None
            try:
                return int(s)
            except ValueError:
                return None

        return None

    def get_double(self, used_row: int, used_col: int) -> Optional[float]:
        """
        Return the cell value as Optional[float] (double precision in Python).

        Conversion rules:
        - int/float => float(value)
        - bool => reject (return None)
        - str => try float(stripped) (supports "3.14", "1e-3", etc.)
        - otherwise => None
        """
        v = self.get_value_by_used_coords(used_row, used_col)

        if isinstance(v, bool):
            return None

        if isinstance(v, (int, float)):
            return float(v)

        if isinstance(v, str):
            s = v.strip()
            if s == "":
                return None
            try:
                return float(s)
            except ValueError:
                return None

        return None

    # -------------------------
    # First / last cell in a used-area row (requested)
    # -------------------------

    def first_cell_in_used_row(self, used_row: int) -> Tuple[int, Any]:
        """
        Return (used_col_index, value) for the first *used* cell in the given used-row.

        "First cell in row" is interpreted as first non-empty value cell within used area.
        If the entire row is empty (possible if used area is a rectangle), returns (0, None).
        """
        area = self.used_area()
        if area.is_empty:
            return (0, None)

        abs_row = area.min_row + (used_row - 1)

        for used_col in range(1, area.ncols + 1):
            abs_col = area.min_col + (used_col - 1)
            v = self.get_value_by_excel_coords(abs_row, abs_col)
            if self._is_value_used(v):
                return (used_col, v)

        return (0, None)

    def last_cell_in_used_row(self, used_row: int) -> Tuple[int, Any]:
        """
        Return (used_col_index, value) for the last *used* cell in the given used-row.

        If the entire row is empty, returns (0, None).
        """
        area = self.used_area()
        if area.is_empty:
            return (0, None)

        abs_row = area.min_row + (used_row - 1)

        for used_col in range(area.ncols, 0, -1):
            abs_col = area.min_col + (used_col - 1)
            v = self.get_value_by_excel_coords(abs_row, abs_col)
            if self._is_value_used(v):
                return (used_col, v)

        return (0, None)


# ---------------------------------------------------------------------------
# Small CLI demo (optional, but handy for quick tests)
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Value-only Excel reader demo.")
    parser.add_argument("xlsx", help="Path to .xlsx file")
    parser.add_argument("sheet_name", nargs="?", default=None, help="Optional sheet name to select")
    parser.add_argument("--sheet-index", type=int, default=None, help="Optional 0-based sheet index to select")
    args = parser.parse_args()

    r = ExcelValueReader(args.xlsx)

    print(f"Sheets: {r.sheet_count()}")
    print("Names :", r.sheet_names())

    # Selection priority: explicit index > explicit name > default first sheet
    if args.sheet_index is not None:
        r.select_sheet_by_index(args.sheet_index)
    elif args.sheet_name is not None:
        r.select_sheet_by_name(args.sheet_name)

    print(f"Selected sheet: {r.selected_sheet_name()}")

    area = r.used_area()
    if area.is_empty:
        print("Used area: (empty — no values found)")
        return

    print(f"Used area bounds: rows {area.min_row}..{area.max_row}, cols {area.min_col}..{area.max_col}")
    print(f"Used area size  : {area.nrows} rows x {area.ncols} cols")

    # Demo: show first and last used cell of the first used-area row
    fc = r.first_cell_in_used_row(1)
    lc = r.last_cell_in_used_row(1)
    print(f"Row 1 first used cell: used_col={fc[0]}, value={fc[1]!r}")
    print(f"Row 1 last  used cell: used_col={lc[0]}, value={lc[1]!r}")

    # Demo: read top-left used-area cell as multiple types
    v_raw = r.get_value_by_used_coords(1, 1)
    print(f"Top-left used cell raw   : {v_raw!r}")
    print(f"Top-left used cell string: {r.get_string(1, 1)!r}")
    print(f"Top-left used cell bool  : {r.get_bool(1, 1)!r}")
    print(f"Top-left used cell int   : {r.get_int(1, 1)!r}")
    print(f"Top-left used cell double: {r.get_double(1, 1)!r}")


if __name__ == "__main__":
    main()
